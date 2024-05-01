
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

movie_data = soup.findAll("tr")



wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'


#write content to a cell


headerfont = Font(size=16, bold=True)

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Number of Theaters'
ws['E1'] = 'Total Gross'
ws['F1'] = 'Average gross by theater'

ws['A1'].font = headerfont
ws['B1'].font = headerfont
ws['C1'].font = headerfont
ws['D1'].font = headerfont
ws['E1'].font = headerfont
ws['F1'].font = headerfont

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 16
ws.column_dimensions['D'].width = 30

row_num = 2

for movie in movie_data[1:6]:
    td=movie.findAll("td")
    rank = td[0].text.strip()
    release_title= td[1].text.strip()
    gross = td[5].text.strip()
    theaters_num = td[6].text.strip()
    total_gross = td[7].text.strip()
    release_date = td[8].text.strip()

    gross_int = int(gross.replace(',', '').replace('$', ''))
    total_gross_int = int(total_gross.replace(',', '').replace('$', ''))

    ws['A' + str(row_num)] = rank
    ws['B'+ str(row_num)] = release_title
    ws['C'+ str(row_num)] = release_date
    ws['D'+ str(row_num)] = theaters_num
    ws['E'+ str(row_num)] = total_gross
    ws['F'+ str(row_num)] = (f" {str(round(gross_int / total_gross_int * 100, 2))}%")
    row_num += 1

wb.save('Box Office Report.xlsx')