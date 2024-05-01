import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

wb.save("PythontoExcel.xlsx")

#write content to a cell

ws['A1'] = 'Invoice'

headerfont = Font(name='Times New Roman', size=24,bold=True)

ws['A1'].font = headerfont

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

#unmerge cellls
#ws.unmerge_cells('A1:B1')

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16, bold=True)

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

# Read the excel file - 'ProduceReport.xlsx' that you created earlier.
# write all the contents of this file to  'Second Sheet' in the current
# workbook

# display the Grand Total and Average of 'Amt Sold' and 'Total'
# at the bottom of the list along with appropriate labels.

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

write_sheet.column_dimensions['A'].width = 25

for row in read_ws.iter_rows():
    ls = [i.value for i in row]

    write_sheet.append(ls)

    max_row = write_sheet.max_row

write_sheet.cell(max_row+2,2).value = 'Total'
write_sheet.cell(max_row+2,2).font = Font(size=16, bold=True)
    
write_sheet.cell(max_row+2,3).value = '=SUM(C2:C' + str(max_row) +')'
write_sheet.cell(max_row+2,4).value = '=SUM(D2:D' + str(max_row) +')'



write_sheet.cell(max_row+4,2).value = 'Average'
write_sheet.cell(max_row+4,2).font = Font(size=16, bold=True) 

write_sheet.cell(max_row+4,3).value = '=AVERAGE(C2:C' + str(max_row) +')'
write_sheet.cell(max_row+4,4).value = '=AVERAGE(D2:D' + str(max_row) +')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'


for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'





'''
i = 1
for currentrow in read_ws.iter_rows(min_row=1, max_row=read_ws.max_row, max_col=read_ws.max_column):
    write_sheet['A' + str(i)] = currentrow[0].value
    write_sheet['B' + str(i)] = currentrow[1].value
    write_sheet['C' + str(i)] = currentrow[2].value
    write_sheet['D' + str(i)] = currentrow[3].value
    i += 1

write_sheet['A'+ str(i+2)] = 'Grand Total'
write_sheet['A'+ str(i+4)] = 'Average'

write_sheet['C'+ str(i+2)] = '=SUM(C2:C41)'
write_sheet['D'+ str(i+2)] = '=SUM(D2:D41)'

write_sheet['C'+ str(i+4)] = '=AVERAGE(C2:C41)'
write_sheet['D'+ str(i+4)] = '=AVERAGE(D2:D41)'

write_sheet['A1'].font = headerfont
write_sheet['A'+ str(i+2)].font = headerfont
write_sheet['A'+ str(i+4)].font = headerfont
'''

wb.save('PythontoExcel.xlsx')